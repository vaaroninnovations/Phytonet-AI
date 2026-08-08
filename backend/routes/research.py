"""AI Research Assistant HTTP router.

Endpoints
─────────
  POST   /api/research/projects             → create a new project
  GET    /api/research/projects             → list current user's projects
  GET    /api/research/projects/{pid}       → project with messages + runs
  DELETE /api/research/projects/{pid}       → hard delete

  POST   /api/research/projects/{pid}/message
      Body: {prompt: str, attachments?: [{name, kind, content_preview}]}
      Returns the planner's response (plan | followup | chat), stores the
      user + assistant messages on the project, and creates a pending run
      when a plan is produced.

  POST   /api/research/projects/{pid}/execute/{run_id}
      Kicks off sequential execution of the pending run. Returns immediately
      with the run doc; the client polls `/status/{run_id}` for progress.

  GET    /api/research/projects/{pid}/status/{run_id}
      Poll a run's status + step results.

  POST   /api/research/projects/{pid}/upload
      Multipart file upload (SMILES / CSV / Excel / MOL / SDF). Returns a
      lightweight descriptor the client can attach to the next message.
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
from datetime import datetime, timezone
from typing import Optional

from bson import ObjectId
from fastapi import (APIRouter, BackgroundTasks, Depends, File, Form,
                     HTTPException, Request, UploadFile)
from pydantic import BaseModel, Field

import auth_service
import research_service

logger = logging.getLogger(__name__)


class NewProjectPayload(BaseModel):
    title: Optional[str] = Field(None, max_length=200)


class MessagePayload(BaseModel):
    prompt: str = Field(..., min_length=1, max_length=6000)
    attachments: list[dict] | None = None


def _serialize_project(d: dict, include_messages: bool = False) -> dict:
    out = {
        "id":        str(d["_id"]),
        "title":     d.get("title") or "New Research",
        "user_id":   d.get("user_id"),
        "created_at": d.get("created_at").isoformat() if d.get("created_at") else None,
        "updated_at": d.get("updated_at").isoformat() if d.get("updated_at") else None,
        "message_count": len(d.get("messages") or []),
        "run_count": len(d.get("runs") or []),
    }
    if include_messages:
        out["messages"] = d.get("messages") or []
        out["runs"] = [_serialize_run(r) for r in (d.get("runs") or [])]
    return out


def _serialize_run(r: dict) -> dict:
    return {
        "id":     r.get("id"),
        "title":  r.get("title"),
        "status": r.get("status", "pending"),
        "plan":   r.get("plan") or [],
        "cost":   r.get("cost"),
        "reasoning": r.get("reasoning"),
        "results": r.get("results") or [],
        "interpretation": r.get("interpretation"),
        "interp_streaming": bool(r.get("interp_streaming", False)),
        "next_steps": r.get("next_steps") or [],
        "created_at": r.get("created_at"),
        "completed_at": r.get("completed_at"),
    }


def build_router(db) -> APIRouter:
    router = APIRouter(prefix="/research", tags=["research"])
    col = db["research_projects"]
    require_user = auth_service.make_get_current_user(db)

    @router.post("/projects")
    async def create_project(payload: NewProjectPayload, user=Depends(require_user)):
        now = datetime.now(timezone.utc)
        doc = {
            "user_id": str(user["_id"]),
            "title":   payload.title or "New Research",
            "messages": [],
            "runs":     [],
            "created_at": now,
            "updated_at": now,
        }
        res = await col.insert_one(doc)
        doc["_id"] = res.inserted_id
        return _serialize_project(doc, include_messages=True)

    @router.get("/projects")
    async def list_projects(user=Depends(require_user)):
        cur = col.find({"user_id": str(user["_id"])}).sort("updated_at", -1)
        return [_serialize_project(d) async for d in cur]

    @router.get("/projects/{pid}")
    async def get_project(pid: str, user=Depends(require_user)):
        d = await _fetch_owned(col, pid, user)
        return _serialize_project(d, include_messages=True)

    @router.delete("/projects/{pid}")
    async def delete_project(pid: str, user=Depends(require_user)):
        d = await _fetch_owned(col, pid, user)
        await col.delete_one({"_id": d["_id"]})
        return {"ok": True}

    @router.post("/projects/{pid}/message")
    async def send_message(pid: str, payload: MessagePayload,
                           user=Depends(require_user)):
        d = await _fetch_owned(col, pid, user)
        now_iso = datetime.now(timezone.utc).isoformat()

        # User message
        user_msg = {
            "role": "user", "text": payload.prompt,
            "attachments": payload.attachments or [],
            "created_at": now_iso,
        }
        history = d.get("messages") or []

        planner = await research_service.plan(
            payload.prompt, history, pid, payload.attachments,
        )
        assistant_msg: dict = {"role": "assistant", "created_at": now_iso,
                               "mode": planner.get("mode", "chat")}
        run_doc: Optional[dict] = None
        title_updates: dict = {}

        if planner.get("mode") == "plan" and planner.get("plan"):
            run_id = str(ObjectId())
            # Compute preflight node cost so the frontend can render
            # "This plan will use ~14 nodes" on the plan card.
            from routes import nodes as _nodes
            cost_breakdown = _nodes.compute_research_run_cost(planner["plan"])
            # Fetch the user's current balance + free-run status
            balance = None; free_runs_used = 0
            try:
                udoc = await db["users"].find_one({"_id": ObjectId(str(user["_id"]))})
                balance = int((udoc or {}).get("nodes_balance") or 0)
                free_runs_used = int((udoc or {}).get("free_research_runs_used") or 0)
            except Exception:
                pass
            free_runs_left = max(0, _nodes.FREE_RESEARCH_RUNS - free_runs_used)
            cost_breakdown["balance"]      = balance
            cost_breakdown["free_runs_left"] = free_runs_left
            cost_breakdown["billable"]     = free_runs_left <= 0
            cost_breakdown["insufficient"] = (free_runs_left <= 0
                                              and balance is not None
                                              and balance < cost_breakdown["total"])
            # ── Analytics: preflight funnel event ──
            try:
                from routes import admin_business as _biz
                await _biz.log_event(
                    db, "research_preflight", str(user["_id"]),
                    module="phytonet-ai-agent",
                    nodes_charged=0,
                    meta={"cost_total": cost_breakdown.get("total"),
                          "docking_pairs": cost_breakdown.get("docking_pairs"),
                          "billable": cost_breakdown.get("billable"),
                          "insufficient": cost_breakdown.get("insufficient")},
                )
            except Exception:
                pass
            run_doc = {
                "id":     run_id,
                "title":  planner.get("title") or "Workflow",
                "status": "pending",
                "reasoning": planner.get("reasoning"),
                "plan":   planner["plan"],
                "cost":   cost_breakdown,
                "results": [],
                "created_at": now_iso,
                "completed_at": None,
            }
            assistant_msg.update({
                "text": planner.get("reasoning") or "Here's the plan I'll run.",
                "run_id": run_id,
                "title":  run_doc["title"],
                "plan":   run_doc["plan"],
                "cost":   cost_breakdown,
            })
            # Auto-set project title on the first plan run
            if (d.get("title") in (None, "", "New Research")
                    and run_doc["title"]):
                title_updates["title"] = run_doc["title"][:120]
        elif planner.get("mode") == "followup":
            assistant_msg["text"] = (planner.get("followup_question")
                                     or "Could you share more detail?")
        else:  # chat
            assistant_msg["text"] = (planner.get("reply")
                                     or "Let me know when you're ready.")

        update = {
            "$push": {"messages": {"$each": [user_msg, assistant_msg]}},
            "$set":  {"updated_at": datetime.now(timezone.utc), **title_updates},
        }
        if run_doc:
            update["$push"]["runs"] = run_doc
        await col.update_one({"_id": d["_id"]}, update)
        return {"user_message": user_msg, "assistant_message": assistant_msg,
                "run": run_doc}

    @router.post("/projects/{pid}/execute/{run_id}")
    async def execute_run(pid: str, run_id: str, bg: BackgroundTasks,
                          user=Depends(require_user)):
        d = await _fetch_owned(col, pid, user)
        run = next((r for r in (d.get("runs") or []) if r.get("id") == run_id),
                   None)
        if not run:
            raise HTTPException(404, "Run not found")
        if run.get("status") == "running":
            raise HTTPException(409, "Run already in progress")
        if run.get("status") == "completed":
            return _serialize_run(run)
        # ── Analytics: research_executed event ──
        try:
            from routes import admin_business as _biz
            cb = run.get("cost") or {}
            await _biz.log_event(
                db, "research_executed", str(user["_id"]),
                module="phytonet-ai-agent",
                nodes_charged=int(cb.get("total") or 0) if cb.get("billable") else 0,
                meta={"run_id": run_id, "billable": cb.get("billable"),
                      "docking_pairs": cb.get("docking_pairs")},
            )
        except Exception:
            pass

        # Mark running
        await col.update_one(
            {"_id": d["_id"], "runs.id": run_id},
            {"$set": {"runs.$.status": "running"}},
        )
        bg.add_task(_execute_in_background, db, str(d["_id"]), pid, run_id)
        run["status"] = "running"
        return _serialize_run(run)

    @router.get("/projects/{pid}/status/{run_id}")
    async def status(pid: str, run_id: str, user=Depends(require_user)):
        d = await _fetch_owned(col, pid, user)
        run = next((r for r in (d.get("runs") or []) if r.get("id") == run_id),
                   None)
        if not run:
            raise HTTPException(404, "Run not found")
        return _serialize_run(run)

    @router.post("/projects/{pid}/upload")
    async def upload(pid: str, file: UploadFile = File(...),
                     user=Depends(require_user)):
        await _fetch_owned(col, pid, user)
        raw = await file.read()
        kind, preview, extracted = _parse_upload(file.filename, raw)
        return {
            "name": file.filename,
            "kind": kind,
            "size": len(raw),
            "content_preview": preview,
            "extracted": extracted,   # e.g. list of SMILES parsed from CSV
        }

    @router.post("/projects/{pid}/share")
    async def enable_share(pid: str, user=Depends(require_user)):
        """Generate (or return existing) public read-only slug for a project."""
        import secrets
        d = await _fetch_owned(col, pid, user)
        slug = d.get("share_slug") or secrets.token_urlsafe(10)
        await col.update_one(
            {"_id": d["_id"]},
            {"$set": {"share_slug": slug, "shared_at":
                      datetime.now(timezone.utc)}},
        )
        return {"share_slug": slug,
                "share_path": f"/research/shared/{slug}"}

    @router.delete("/projects/{pid}/share")
    async def disable_share(pid: str, user=Depends(require_user)):
        d = await _fetch_owned(col, pid, user)
        await col.update_one({"_id": d["_id"]},
                             {"$unset": {"share_slug": "", "shared_at": ""}})
        return {"ok": True}

    @router.post("/projects/{pid}/retry/{run_id}/{step_id}")
    async def retry_step(pid: str, run_id: str, step_id: str,
                          background_tasks: BackgroundTasks,
                          user=Depends(require_user)):
        """Reset a failed step (+ every downstream step in the plan) back to
        'pending', clear their results, and re-execute the run. Previously-
        completed steps upstream are preserved untouched."""
        d = await _fetch_owned(col, pid, user)
        run = next((r for r in (d.get("runs") or []) if r.get("id") == run_id),
                   None)
        if not run:
            raise HTTPException(404, "Run not found")
        plan = run.get("plan") or []
        target_idx = next((i for i, s in enumerate(plan)
                           if s.get("id") == step_id), None)
        if target_idx is None:
            raise HTTPException(404, "Step not found in plan")

        # Reset this step + all downstream steps back to `pending`.
        new_plan = []
        for i, s in enumerate(plan):
            if i >= target_idx:
                new_plan.append({**s, "status": "pending", "progress": None})
            else:
                new_plan.append(s)

        # Trim results to only steps strictly before the retried step.
        old_results = run.get("results") or []
        new_results = [r for i, r in enumerate(old_results) if i < target_idx]

        await col.update_one(
            {"_id": d["_id"], "runs.id": run_id},
            {"$set": {
                "runs.$.plan":           new_plan,
                "runs.$.results":        new_results,
                "runs.$.status":         "running",
                "runs.$.interpretation": "",
                "runs.$.next_steps":     [],
                "runs.$.completed_at":   None,
            }},
        )
        background_tasks.add_task(_execute_in_background,
                                   db, str(d["_id"]), pid, run_id)
        return {"ok": True, "retried_from": step_id,
                "reset_steps": len(plan) - target_idx}

    return router


def build_public_share_router(db) -> APIRouter:
    """Unauthenticated read-only access to shared projects."""
    router = APIRouter(prefix="/research", tags=["research-public"])
    col = db["research_projects"]

    @router.get("/shared/{slug}")
    async def get_shared(slug: str):
        d = await col.find_one({"share_slug": slug})
        if not d:
            raise HTTPException(404, "Shared project not found")
        # Redact user_id + IPs before returning
        return {
            "id":        str(d["_id"]),
            "title":     d.get("title"),
            "shared_at": d.get("shared_at").isoformat() if d.get("shared_at") else None,
            "messages":  d.get("messages") or [],
            "runs":     [_serialize_run(r) for r in (d.get("runs") or [])],
        }

    return router


# ─────────────────────── helpers ─────────────────────────────────
async def _fetch_owned(col, pid: str, user) -> dict:
    try:
        oid = ObjectId(pid)
    except Exception:
        raise HTTPException(400, "Invalid project id")
    d = await col.find_one({"_id": oid, "user_id": str(user["_id"])})
    if not d:
        raise HTTPException(404, "Project not found")
    return d


async def _execute_in_background(db, oid_str: str, pid: str, run_id: str):
    """Sequential executor. Streams progress by mutating the run doc as each
    step completes, so the client's poller can render live progress.

    If the run already has partial `results` (e.g. re-invocation after a
    retry), previously-`done` steps are preserved and only pending / error
    steps are re-executed."""
    col = db["research_projects"]
    oid = ObjectId(oid_str)
    d = await col.find_one({"_id": oid})
    run = next(r for r in d.get("runs", []) if r.get("id") == run_id)
    plan_steps = run.get("plan") or []
    # Project owner — used for node metering per-tool.
    owner_user_id = d.get("user_id")

    # Cross-run context — flat list of every step from earlier COMPLETED runs.
    # Placeholder resolution + admet_predict auto-injection use this when the
    # current run's plan doesn't include a producing step.
    project_context: list[dict] = []
    for prior_run in (d.get("runs") or []):
        if prior_run.get("id") == run_id:
            break
        if prior_run.get("status") != "completed":
            continue
        for r in (prior_run.get("results") or []):
            if r.get("status") == "done":
                project_context.append(r)

    # Also treat uploaded attachments as pseudo-steps so ADMET can auto-inject
    # SMILES from any CSV / SMI / XLSX the user has attached in this project.
    for msg in (d.get("messages") or []):
        for att in (msg.get("attachments") or []):
            smis = att.get("extracted") or []
            if isinstance(smis, list) and smis:
                project_context.append({
                    "id":     f"attachment_{att.get('name','file')}",
                    "status": "done",
                    "tool":   "user_attachment",
                    "result": {"status": "ok",
                               "card":   "compound_table",
                               "data":   {
                                   "smiles_extracted": smis,
                                   "compounds": [{"smiles": s} for s in smis],
                               }},
                })

    # Preserve already-done step results so retries only rerun what failed.
    prior_step_results: dict[str, dict] = {}
    for r in (run.get("results") or []):
        if r.get("status") == "done" and r.get("id"):
            prior_step_results[r["id"]] = r

    results: list[dict] = []
    for idx, step in enumerate(plan_steps):
        step_id = step.get("id")
        # Skip steps that already succeeded — reuse their prior result.
        if step_id in prior_step_results and step.get("status") == "done":
            results.append(prior_step_results[step_id])
            continue
        # Mark running on this step
        await col.update_one(
            {"_id": oid, "runs.id": run_id},
            {"$set": {
                f"runs.$.plan.{idx}.status":   "running",
                f"runs.$.plan.{idx}.progress": {"stage": "starting",
                                                "detail": "Starting…"},
            }},
        )

        # Progress callback — updates Mongo so the frontend poller sees
        # live sub-status (e.g. "Querying IMPPAT…", "Removing duplicates…").
        # Optional `partial` payload lets long-running tools (docking!)
        # persist intermediate results so the frontend can render a live-
        # partial card instead of waiting for the full batch to finish.
        async def _step_progress(stage: str, detail: str = "",
                                   partial: dict | None = None,
                                   _idx=idx):
            try:
                update = {
                    f"runs.$.plan.{_idx}.progress": {
                        "stage": stage,
                        "detail": detail,
                        "at": datetime.now(timezone.utc).isoformat(),
                    }
                }
                if partial is not None:
                    update[f"runs.$.plan.{_idx}.partial_result"] = partial
                await col.update_one(
                    {"_id": oid, "runs.id": run_id},
                    {"$set": update},
                )
            except Exception as e:
                logger.warning(f"[research] progress persist failed: {e}")

        result = await research_service.execute_step(
            step, prior_results=results, project_context=project_context,
            progress=_step_progress,
        )
        status = "done" if result.get("status") == "ok" else "error"

        # ── Node metering — charge on success only ──────────────
        # Free-tier grace covers the first FREE_RESEARCH_RUNS of every
        # user (see routes/nodes.py). Beyond that the user is billed
        # per-tool with an idempotent debit keyed on (run_id, step_id).
        charge_info = None
        if status == "done":
            from routes import nodes as _nodes
            if step.get("tool") == "docking":
                pairs = ((result.get("data") or {}).get("metrics") or {}).get("n_pairs")
                if not pairs:
                    args = step.get("args") or {}
                    tc = int(args.get("top_compounds") or 5)
                    tg = int(args.get("top_genes")     or 3)
                    pairs = tc * tg
                cost = int(pairs) * _nodes.RESEARCH_DOCKING_COST_PER_PAIR
            else:
                cost = _nodes.RESEARCH_TOOL_COSTS.get(step.get("tool"), 0)
            if cost > 0:
                charge_info = await _nodes.research_charge_step(
                    user_id=owner_user_id or "",
                    run_id=run_id, step_id=step_id,
                    tool=step.get("tool"), amount=cost,
                )
                logger.info(f"[research] charge {step.get('tool')} "
                            f"({cost} nodes) → {charge_info}")

        step_out = {
            "id": step_id,
            "label": step.get("label"),
            "tool": step.get("tool"),
            "args": step.get("args"),
            "status": status,
            "result": result,
            "cost": charge_info,
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }
        results.append(step_out)
        await col.update_one(
            {"_id": oid, "runs.id": run_id},
            {"$set": {
                f"runs.$.plan.{idx}.status":   status,
                f"runs.$.plan.{idx}.progress": {
                    "stage":  "completed" if status == "done" else "failed",
                    "detail": (result.get("message")
                               if status == "done"
                               else result.get("message") or "Failed"),
                    "at": datetime.now(timezone.utc).isoformat(),
                },
                "runs.$.results": results,
            }},
        )
        # Stop on hard error
        if status == "error":
            break

    # Stream the natural-language interpretation token-by-token, saving
    # partial text to the run doc as each chunk arrives. Existing /status
    # polling then surfaces the growing text progressively — no SSE needed.
    interpretation = ""
    next_steps: list[str] = []
    try:
        buf: list[str] = []
        # Flush to Mongo every N chars (or every 300 ms) to avoid write storms.
        FLUSH_EVERY = 40
        last_flush_len = 0
        async for delta in research_service.interpret_stream(
                {"title": run.get("title"),
                 "reasoning": run.get("reasoning"),
                 "plan": plan_steps},
                results, pid):
            if not delta:
                continue
            buf.append(delta)
            joined_len = sum(len(x) for x in buf)
            if joined_len - last_flush_len >= FLUSH_EVERY:
                last_flush_len = joined_len
                await col.update_one(
                    {"_id": oid, "runs.id": run_id},
                    {"$set": {
                        "runs.$.interpretation": "".join(buf),
                        "runs.$.interp_streaming": True,
                    }},
                )
        interpretation = "".join(buf).strip()
    except Exception as e:
        logger.warning(f"[research] interpret error: {e}")
    try:
        next_steps = await research_service.suggest_next_steps(
            {"title": run.get("title"), "plan": plan_steps},
            results, pid,
        )
    except Exception as e:
        logger.warning(f"[research] next_steps error: {e}")

    final_status = "completed" if all(
        r["status"] == "done" for r in results) else "failed"
    now_iso = datetime.now(timezone.utc).isoformat()
    await col.update_one(
        {"_id": oid, "runs.id": run_id},
        {"$set": {
            "runs.$.status":            final_status,
            "runs.$.interpretation":    interpretation,
            "runs.$.interp_streaming":  False,
            "runs.$.next_steps":        next_steps,
            "runs.$.completed_at":      now_iso,
        }},
    )
    # Append the interpretation as an assistant message so it stays in chat
    if interpretation:
        await col.update_one(
            {"_id": oid},
            {"$push": {"messages": {
                "role": "assistant",
                "text": interpretation,
                "mode": "interpretation",
                "run_id": run_id,
                "next_steps": next_steps,
                "created_at": now_iso,
            }},
             "$set": {"updated_at": datetime.now(timezone.utc)}},
        )


def _parse_upload(name: str, raw: bytes) -> tuple[str, str, list]:
    """Return (kind, preview_text, extracted_list). Extracts SMILES from
    CSV/TXT/XLSX where possible."""
    name_lc = (name or "").lower()
    text = ""
    try:
        text = raw.decode("utf-8", errors="ignore")
    except Exception:
        pass

    extracted: list[str] = []
    kind = "unknown"

    if name_lc.endswith(".smi") or name_lc.endswith(".txt"):
        kind = "smiles"
        for line in text.splitlines():
            s = line.strip().split()
            if s and _looks_like_smiles(s[0]):
                extracted.append(s[0])
    elif name_lc.endswith(".csv"):
        kind = "csv"
        try:
            reader = csv.reader(io.StringIO(text))
            header = next(reader, [])
            smi_idx = next((i for i, h in enumerate(header)
                            if h.strip().lower() in ("smiles", "canonical_smiles")),
                           None)
            for row in reader:
                if smi_idx is not None and smi_idx < len(row):
                    if _looks_like_smiles(row[smi_idx]):
                        extracted.append(row[smi_idx].strip())
                elif row and _looks_like_smiles(row[0]):
                    extracted.append(row[0].strip())
        except Exception:
            pass
    elif name_lc.endswith((".xlsx", ".xls")):
        kind = "excel"
        try:
            import openpyxl  # lazy — only when needed
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            ws = wb.active
            header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
            smi_idx = next((i for i, h in enumerate(header)
                            if h in ("smiles", "canonical_smiles")), None)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                cand = str(row[smi_idx] if smi_idx is not None else row[0] or "")
                if _looks_like_smiles(cand):
                    extracted.append(cand.strip())
        except Exception as e:
            logger.warning(f"excel parse failed: {e}")
    elif name_lc.endswith((".mol", ".sdf")):
        kind = "sdf"  # backend/openbabel handles it downstream; we keep raw
    elif text:
        kind = "text"

    preview = text[:400] if text else f"<binary {len(raw)} bytes>"
    return kind, preview, extracted[:250]


def _looks_like_smiles(s: str) -> bool:
    s = (s or "").strip()
    if not (3 <= len(s) <= 300):
        return False
    # Cheap heuristic: contains at least one letter typical of SMILES
    return any(c.isalpha() for c in s) and any(c in s for c in "CNOSFPBH()[]=#")
